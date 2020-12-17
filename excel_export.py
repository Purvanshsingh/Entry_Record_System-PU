from openpyxl import Workbook,load_workbook
from datetime import date
import os
def create_excel_sheet(data):
    today = date.today()
    filename = today.strftime("%b-%d-%Y")
    filename = filename + '.xlsx'
    file_location = os.path.join('Entries',filename)
    print(file_location)
    try:
        workbook = load_workbook(file_location)
        sheet=workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "REGISTRATION NO"
        sheet['B1'] = "TIME"
        sheet['C1'] = "DESIGNATION"

    sheet.append(data)
    workbook.save(filename=file_location)